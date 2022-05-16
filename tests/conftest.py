import pytest
from appium import webdriver
'''
import openpyxl

Dict={}
book =openpyxl.load_workbook("C:\\Users\\karthiks\\Desktop\\package_name.xlsx")
sheet=book.active
Dict[sheet.cell(row=1,column=2).value]=sheet.cell(row=2,column=2).value
Dict[sheet.cell(row=1,column=3).value]=sheet.cell(row=2,column=3).value
'''
@pytest.fixture()
def setup(request):
    desired_caps = {
        "appPackage": 'com.google.android.contacts',
        "appActivity": 'com.android.contacts.activities.PeopleActivity',
        "automationName": "uiautomator2",
        "platformName": "Android",
        "deviceName": "POCO F1",
        "ignoreHiddenApiPolicyError":True,
        'autoGrantPermissions':True
    }
    driver = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)
    request.cls.driver=driver
    yield
    driver.close_app()